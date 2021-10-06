

#SURESH SAHU
#1901CE52
#TUT5

import openpyxl
import os
try:
    os.makedirs("output")
except FileExistsError:
    pass

import csv

with open("names-roll.csv", 'r') as fldrf:
   rows=csv.reader(fldrf)
   sgsg={} 
   for i in rows:
      if(i[0]=="Roll"):continue
      sgsg[i[0]]=f"{i[1].strip()}"

with open("subjects_master.csv" , 'r') as fldrf:
    rows=csv.reader(fldrf)
    subdic={}
    for i in rows:
        if i[0]== 'subno':continue
        subdic[i[0]]=[i[1],i[2]]

with open("grades.csv",'r') as fldrf:
    rows = csv.reader(fldrf)
    for i in rows:
        if i[0] == 'Roll':continue
        with open(fr"output\{i[0]}_{i[1]}.csv" , "a" , newline='') as fldrf:
            writer=csv.writer(fldrf)
            writer.writerow([i[2],subdic[i[2]][0] ,subdic[i[2]][1] ,i[3] ,i[5] , i[4]])

grde={"AA":"10","AB":"9","BB":"8","BC":"7","CC":"6","CD":"5","DD":"4","F":"0","I":"0","F*":"0","I*":"0","DD*":"4"," BB":"8"}
seme1cr=  { a:0 for a in sgsg}
seme2cr=  { a:0 for a in sgsg}
seme3cr=  { a:0 for a in sgsg}
seme4cr=  { a:0 for a in sgsg}
seme5cr=  { a:0 for a in sgsg}
seme6cr=  { a:0 for a in sgsg}
seme7cr=  { a:0 for a in sgsg}
seme8cr=  { a:0 for a in sgsg}
seme10cr= { a:0 for a in sgsg}
seme1u=  { a:0 for a in sgsg}
seme2u=  { a:0 for a in sgsg}
seme3u=  { a:0 for a in sgsg}
seme4u=  { a:0 for a in sgsg}
seme5u=  { a:0 for a in sgsg}
seme6u=  { a:0 for a in sgsg}
seme7u=  { a:0 for a in sgsg}
seme8u=  { a:0 for a in sgsg}
seme10u=  { a:0 for a in sgsg}

from openpyxl import Workbook
from openpyxl import load_workbook
for x,y in sgsg.items():
    
    wb = Workbook()    
    wb.remove(wb["Sheet"])  
    from pathlib import Path
    config=Path(fr"output\{x}_1.csv")
    if config.is_file():       
      wb.create_sheet(index=1, title="Sem1")  
      Sem1=wb["Sem1"]
      Sem1.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_1.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme1cr[x]=int(seme1cr[x])+int(i[3])
            seme1u[x]=int(seme1u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem1.max_row
            i.insert(0, row_count)           
            Sem1.append(i)
      os.remove(fr"output\{x}_1.csv") 
    from pathlib import Path
    config=Path(fr"output\{x}_2.csv")
    if config.is_file():     
      wb.create_sheet(index=2, title="Sem2")  
      Sem2=wb["Sem2"]
      Sem2.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      from openpyxl.styles import Font
      from openpyxl.styles import PatternFill
      for row in Sem2.iter_rows(min_row=1, max_row=1):
          for cell in row:
              cell.font = Font(bold=True)
              cell.fill = PatternFill(fgColor="77C3FD", fill_type = "solid")
      with open(fr"output\{x}_2.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme2cr[x]=int(seme2cr[x])+int(i[3])
            seme2u[x]=int(seme2u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem2.max_row
            i.insert(0, row_count)           
            Sem2.append(i)
        for row in Sem2.iter_rows(min_row=2,max_row=row_count+1):
           for cell in row:
              cell.font = Font(bold=True)
              cell.fill = PatternFill(fgColor="B5DDFB", fill_type = "solid")         
      os.remove(fr"output\{x}_2.csv")  
    from pathlib import Path
    config=Path(fr"output\{x}_3.csv")
    if config.is_file():                   
      wb.create_sheet(index=3, title="Sem3")  
      Sem3=wb["Sem3"]
      Sem3.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_3.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme3cr[x]=int(seme3cr[x])+int(i[3])
            seme3u[x]=int(seme3u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem3.max_row
            i.insert(0, row_count)           
            Sem3.append(i)
      os.remove(fr"output\{x}_3.csv")
    from pathlib import Path
    config=Path(fr"output\{x}_4.csv")
    if config.is_file():         
      wb.create_sheet(index=4, title="Sem4")  
      Sem4=wb["Sem4"]
      Sem4.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_4.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme4cr[x]=int(seme4cr[x])+int(i[3])
            seme4u[x]=int(seme4u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem4.max_row
            i.insert(0, row_count)           
            Sem4.append(i)
      os.remove(fr"output\{x}_4.csv")
    from pathlib import Path
    config=Path(fr"output\{x}_5.csv")
    if config.is_file():         
      wb.create_sheet(index=5, title="Sem5")  
      Sem5=wb["Sem5"]
      Sem5.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_5.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme5cr[x]=int(seme5cr[x])+int(i[3])
            seme5u[x]=int(seme5u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem5.max_row
            i.insert(0, row_count)           
            Sem5.append(i)  
      os.remove(fr"output\{x}_5.csv")  
    from pathlib import Path
    config=Path(fr"output\{x}_6.csv")
    if config.is_file():                            
      wb.create_sheet(index=6, title="Sem6")  
      Sem6=wb["Sem6"]
      Sem6.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_6.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme6cr[x]=int(seme6cr[x])+int(i[3])
            seme6u[x]=int(seme6u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem6.max_row
            i.insert(0, row_count)           
            Sem6.append(i)           
      os.remove(fr"output\{x}_6.csv")  
    from pathlib import Path
    config=Path(fr"output\{x}_7.csv")
    if config.is_file():       
      wb.create_sheet(index=7, title="Sem7")  
      Sem7=wb["Sem7"]
      Sem7.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_7.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme7cr[x]=int(seme7cr[x])+int(i[3])
            seme7u[x]=int(seme7u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem7.max_row
            i.insert(0, row_count)           
            Sem7.append(i)  
      os.remove(fr"output\{x}_7.csv")   
    from pathlib import Path
    config=Path(fr"output\{x}_8.csv")
    if config.is_file():                
      wb.create_sheet(index=8, title="Sem8")  
      Sem8=wb["Sem8"]
      Sem8.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_8.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme8cr[x]=int(seme8cr[x])+int(i[3])
            seme8u[x]=int(seme8u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem8.max_row
            i.insert(0, row_count)           
            Sem8.append(i)
      os.remove(fr"output\{x}_8.csv")        
    from pathlib import Path
    config=Path(fr"output\{x}_10.csv")
    if config.is_file(): 
      wb.create_sheet(index=10, title="Sem10")  
      Sem10=wb["Sem10"]
      Sem10.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(fr"output\{x}_10.csv", 'r') as fldrf:
        rows=csv.reader(fldrf)
        for i in rows:
            seme10cr[x]=int(seme10cr[x])+int(i[3])
            seme10u[x]=int(seme10u[x])+int(i[3])*int(grde[i[5]])
            row_count = Sem10.max_row
            i.insert(0, row_count)           
            Sem10.append(i)
      os.remove(fr"output\{x}_10.csv")        

    wb.create_sheet(index=0, title="Overall")  
    Overall=wb["Overall"]
    if(seme3cr[x]==0):
       Overall.append(["Roll No.",f"{x}"])
       Overall.append(["Name of Student",f"{y}"])
       Overall.append(["Discipi",f"{x[4:6]}"])
       Overall.append(["Semester No.","1","2"])
       Overall.append(["Semester wise Credit Taken",f"{seme1cr[x]}",f"{seme2cr[x]}"])
       Overall.append(["SPI",str(round(seme1u[x]/seme1cr[x],2)),str(round(seme2u[x]/seme2cr[x],2))])
       Overall.append(["Total Credits Taken",f"{seme1cr[x]}",f"{seme1cr[x]+seme2cr[x]}"])
       Overall.append(["CPI",str(round(seme1u[x]/seme1cr[x],2)),str(round((seme1u[x]+seme2u[x])/(seme1cr[x]+seme2cr[x]),2))])
    elif(seme5cr[x]==0):
        Overall.append(["Roll No.",f"{x}"])
        Overall.append(["Name of Student",f"{y}"])
        Overall.append(["Discipi",f"{x[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4"])
        Overall.append(["Semester wise Credit Taken",f"{seme1cr[x]}",f"{seme2cr[x]}",f"{seme3cr[x]}",f"{seme4cr[x]}"])
        Overall.append(["SPI",str(round(seme1u[x]/seme1cr[x],2)),str(round(seme2u[x]/seme2cr[x],2)),str(round(seme3u[x]/seme3cr[x],2)),str(round(seme4u[x]/seme4cr[x],2))])
        Overall.append(["Total Credits Taken",f"{seme1cr[x]}",f"{seme1cr[x]+seme2cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]}"])
        Overall.append(["CPI",str(round(seme1u[x]/seme1cr[x],2)),str(round((seme1u[x]+seme2u[x])/(seme1cr[x]+seme2cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]),2))])
    elif(seme6cr[x]==0): 
        Overall.append(["Roll No.",f"{x}"])
        Overall.append(["Name of Student",f"{y}"])
        Overall.append(["Discipi",f"{x[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5"])
        Overall.append(["Semester wise Credit Taken",f"{seme1cr[x]}",f"{seme2cr[x]}",f"{seme3cr[x]}",f"{seme4cr[x]}",f"{seme5cr[x]}"])
        Overall.append(["SPI",str(round(seme1u[x]/seme1cr[x],2)),str(round(seme2u[x]/seme2cr[x],2)),str(round(seme3u[x]/seme3cr[x],2)),str(round(seme4u[x]/seme4cr[x],2)),str(round(seme5u[x]/seme5cr[x],2))])
        Overall.append(["Total Credits Taken",f"{seme1cr[x]}",f"{seme1cr[x]+seme2cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]}"])
        Overall.append(["CPI",str(round(seme1u[x]/seme1cr[x],2)),str(round((seme1u[x]+seme2u[x])/(seme1cr[x]+seme2cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]),2))])
    elif(seme8cr[x]==0):
        Overall.append(["Roll No.",f"{x}"])
        Overall.append(["Name of Student",f"{y}"])
        Overall.append(["Discipi",f"{x[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5","6","7"])
        Overall.append(["Semester wise Credit Taken",f"{seme1cr[x]}",f"{seme2cr[x]}",f"{seme3cr[x]}",f"{seme4cr[x]}",f"{seme5cr[x]}",f"{seme6cr[x]}",f"{seme7cr[x]}"])
        Overall.append(["SPI",str(round(seme1u[x]/seme1cr[x],2)),str(round(seme2u[x]/seme2cr[x],2)),str(round(seme3u[x]/seme3cr[x],2)),str(round(seme4u[x]/seme4cr[x],2)),str(round(seme5u[x]/seme5cr[x],2)),str(round(seme6u[x]/seme6cr[x],2)),str(round(seme7u[x]/seme7cr[x],2))])
        Overall.append(["Total Credits Taken",f"{seme1cr[x]}",f"{seme1cr[x]+seme2cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]}"])
        Overall.append(["CPI",str(round(seme1u[x]/seme1cr[x],2)),str(round((seme1u[x]+seme2u[x])/(seme1cr[x]+seme2cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x]+seme7u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]),2))])
    elif(seme10cr[x]==0):
        Overall.append(["Roll No.",f"{x}"])
        Overall.append(["Name of Student",f"{y}"])
        Overall.append(["Discipi",f"{x[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5","6","7","8"])
        Overall.append(["Semester wise Credit Taken",f"{seme1cr[x]}",f"{seme2cr[x]}",f"{seme3cr[x]}",f"{seme4cr[x]}",f"{seme5cr[x]}",f"{seme6cr[x]}",f"{seme7cr[x]}",f"{seme8cr[x]}"])
        Overall.append(["SPI",str(round(seme1u[x]/seme1cr[x],2)),str(round(seme2u[x]/seme2cr[x],2)),str(round(seme3u[x]/seme3cr[x],2)),str(round(seme4u[x]/seme4cr[x],2)),str(round(seme5u[x]/seme5cr[x],2)),str(round(seme6u[x]/seme6cr[x],2)),str(round(seme7u[x]/seme7cr[x],2)),str(round(seme8u[x]/seme8cr[x],2))])
        Overall.append(["Total Credits Taken",f"{seme1cr[x]}",f"{seme1cr[x]+seme2cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]+seme8cr[x]}"])
        Overall.append(["CPI",str(round(seme1u[x]/seme1cr[x],2)),str(round((seme1u[x]+seme2u[x])/(seme1cr[x]+seme2cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x]+seme7u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x]+seme7u[x]+seme8u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]+seme8cr[x]),2))])
    else: 
        Overall=wb["Overall"] 
        Overall.append(["Roll No.",f"{x}"])
        Overall.append(["Name of Student",f"{y}"])
        Overall.append(["Discipi",f"{x[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5","6","7","8","10"])
        Overall.append(["Semester wise Credit Taken",f"{seme1cr[x]}",f"{seme2cr[x]}",f"{seme3cr[x]}",f"{seme4cr[x]}",f"{seme5cr[x]}",f"{seme6cr[x]}",f"{seme7cr[x]}",f"{seme8cr[x]}",f"{seme10cr[x]}"])
        Overall.append(["SPI",str(round(seme1u[x]/seme1cr[x],2)),str(round(seme2u[x]/seme2cr[x],2)),str(round(seme3u[x]/seme3cr[x],2)),str(round(seme4u[x]/seme4cr[x],2)),str(round(seme5u[x]/seme5cr[x],2)),str(round(seme6u[x]/seme6cr[x],2)),str(round(seme7u[x]/seme7cr[x],2)),str(round(seme8u[x]/seme8cr[x],2)),str(round(seme10u[x]/seme10cr[x],2))])
        Overall.append(["Total Credits Taken",f"{seme1cr[x]}",f"{seme1cr[x]+seme2cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]+seme8cr[x]}",f"{seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]+seme8cr[x]+seme10cr[x]}"])
        Overall.append(["CPI",str(round(seme1u[x]/seme1cr[x],2)),str(round((seme1u[x]+seme2u[x])/(seme1cr[x]+seme2cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x]+seme7u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x]+seme7u[x]+seme8u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]+seme8cr[x]),2)),str(round((seme1u[x]+seme2u[x]+seme3u[x]+seme4u[x]+seme5u[x]+seme6u[x]+seme7u[x]+seme8u[x]+seme10u[x])/(seme1cr[x]+seme2cr[x]+seme3cr[x]+seme4cr[x]+seme5cr[x]+seme6cr[x]+seme7cr[x]+seme8cr[x]+seme10cr[x]),2))])
    
    wb.save(f"output//{x}.xlsx")





