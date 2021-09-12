#1901ce52 SURESH SAHU
# please wait 300 seconds to finish the code to run completely
import os
os.system("cls")

def output_by_subject():
    import csv
    import os
    try:
        os.mkdir("output_by_subject")
    except FileExistsError:
        pass
    with open("regtable_old.csv", "r") as f:
        reader = csv.reader(f)
        for Row in reader:
            del Row[2]
            del Row[3:7]
            if (Row[0] == "rollno"):
                continue
            if(os.path.exists(f"output_by_subject\\{Row[2]}.xlsx")):
                from openpyxl import load_workbook
                wb = load_workbook(f"output_by_subject\\{Row[2]}.xlsx")
                sheet1 = wb.active
                sheet1.append(Row)
                wb.save(f"output_by_subject\\{Row[2]}.xlsx")
            else:
                from openpyxl import Workbook
                wb = Workbook()
                sheet1 = wb.active
                sheet1.append(['rollno','register_sem','subno','sub_type'])
                sheet1.append(Row)
                wb.save(f"output_by_subject\\{Row[2]}.xlsx")
        return


def output_individual_roll():
    import csv
    import os
    try:
        os.mkdir("output_individual_roll")
    except FileExistsError:
        pass
    with open("regtable_old.csv", "r") as f:
        reader = csv.reader(f)
        for Row in reader:
            del Row[2]
            del Row[3:7]
            if (Row[0] == "rollno"):
                continue
            if(os.path.exists(f"output_individual_roll\\{Row[0]}.xlsx")):
                from openpyxl import load_workbook
                wb = load_workbook(f"output_individual_roll\\{Row[0]}.xlsx")
                sheet1 = wb.active
                sheet1.append(Row)
                wb.save(f"output_individual_roll\\{Row[0]}.xlsx")
            else:
                from openpyxl import Workbook
                wb = Workbook()
                sheet1 = wb.active
                sheet1.append(['rollno','register_sem','subno','sub_type'])
                sheet1.append(Row)
                wb.save(f"output_individual_roll\\{Row[0]}.xlsx")
        return


output_by_subject()
output_individual_roll()